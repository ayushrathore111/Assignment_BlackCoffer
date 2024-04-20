import openpyxl
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import cmudict
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from textstat import flesch_reading_ease, syllable_count

nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
nltk.download('vader_lexicon')
nltk.download('cmudict')

def get_pos_neg_scores(text):
    sid = SentimentIntensityAnalyzer()
    scores = sid.polarity_scores(text)
    return scores['pos'], scores['neg']

def get_polarity_subjectivity_scores(text):
    sia = SentimentIntensityAnalyzer()
    polarity_score = sia.polarity_scores(text)['compound']
    subjectivity_score = abs(polarity_score)  # Absolute value of compound score can represent subjectivity
    return polarity_score, subjectivity_score

def get_avg_sentence_length(text):
    sentences = sent_tokenize(text)
    total_words = sum(len(word_tokenize(sentence)) for sentence in sentences)
    return total_words / len(sentences)

def get_percentage_complex_words(text):
    words = word_tokenize(text)
    complex_word_count = sum(1 for word in words if syllable_count(word) > 2)  # Syllable count > 2 indicates a complex word
    return (complex_word_count / len(words)) * 100

def get_fog_index(text):
    words = word_tokenize(text)
    syllable_count_total = sum(syllable_count(word) for word in words)
    avg_sentence_length = get_avg_sentence_length(text)
    return 0.4 * (avg_sentence_length + get_percentage_complex_words(text))

def get_avg_words_per_sentence(text):
    words = word_tokenize(text)
    sentences = sent_tokenize(text)
    return len(words) / len(sentences)

def get_complex_word_count(text):
    words = word_tokenize(text)
    return sum(1 for word in words if syllable_count(word) > 2)

def get_word_count(text):
    return len(word_tokenize(text))

def get_syllables_per_word(text):
    words = word_tokenize(text)
    syllable_count_total = sum(syllable_count(word) for word in words)
    return syllable_count_total / len(words)

def get_personal_pronouns(text):
    words = word_tokenize(text)
    pos_tags = nltk.pos_tag(words)
    pronouns = [word for word, pos in pos_tags if pos == 'PRP']
    return len(pronouns)

def get_avg_word_length(text):
    words = word_tokenize(text)
    total_word_length = sum(len(word) for word in words)
    return total_word_length / len(words)

def analyze_text(text):
    positive_score, negative_score = get_pos_neg_scores(text)
    polarity_score, subjectivity_score = get_polarity_subjectivity_scores(text)
    avg_sentence_length = get_avg_sentence_length(text)
    percentage_complex_words = get_percentage_complex_words(text)
    fog_index = get_fog_index(text)
    avg_words_per_sentence = get_avg_words_per_sentence(text)
    complex_word_count = get_complex_word_count(text)
    word_count = get_word_count(text)
    syllable_per_word = get_syllables_per_word(text)
    personal_pronouns = get_personal_pronouns(text)
    avg_word_length = get_avg_word_length(text)

    return [positive_score, negative_score, polarity_score, subjectivity_score,
            avg_sentence_length, percentage_complex_words, fog_index,
            avg_words_per_sentence, complex_word_count, word_count,
            syllable_per_word, personal_pronouns, avg_word_length]

def main():
    # Load input data
    input_wb = openpyxl.load_workbook('/home/proayush/Desktop/blackoffer/Input.xlsx')
    input_sheet = input_wb.active

    # Create output workbook
    output_wb = openpyxl.load_workbook('/home/proayush/Desktop/blackoffer/Output Data Structure.xlsx')
    output_sheet = output_wb.active

    # Process each row
    for idx, row in enumerate(input_sheet.iter_rows(min_row=2, values_only=True), start=2):
        url_id, url = row
        with open(f"/home/proayush/Desktop/blackoffer/titles_extraction/{url_id}.txt", 'r', encoding='utf-8') as file:
            text = file.read()
            analysis_result = analyze_text(text)
            output_sheet.append([url_id, url] + analysis_result)

    # Save output workbook
    output_wb.save('output.xlsx')

if __name__ == "__main__":
    main()
